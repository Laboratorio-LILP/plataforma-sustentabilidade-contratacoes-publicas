"""Gera a planilha-modelo PESCP — seeds/pescp_seed_template.xlsx.

Não popula com dados reais; apenas cria as 7 abas com seus cabeçalhos
canônicos e uma linha de exemplo comentada na primeira aba (objetos)
para orientar o time de conteúdo.

Uso:
    python tools/gerar_seed_template.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = {
    "objetos_contratuais": [
        "nome", "slug", "descricao", "supercategoria_slug", "ordem",
    ],
    "normas_legais": [
        "titulo_curto", "titulo_completo", "slug", "tipo", "esfera",
        "numero", "data_publicacao", "link_oficial", "ementa",
    ],
    "selos_certificacao": [
        "nome", "slug", "tipo", "descricao", "link_oficial",
    ],
    "tags": [
        "nome", "slug",
    ],
    "criterios": [
        "codigo", "titulo", "slug", "objeto_contratual_slug",
        "tipo_contratacao", "fase_processo", "nivel_ambicao",
        "texto_providencia", "determinacoes_principais",
        "justificativa_tecnica", "metodo_verificacao", "precaucoes",
        "ods_numeros", "normas_slugs", "selos_slugs", "tags_slugs",
        "fonte", "status",
    ],
    "paginas_institucionais": [
        "titulo", "slug", "conteudo", "ordem", "publicada",
    ],
    "glossario": [
        "termo", "slug", "sigla", "definicao",
    ],
}


EXEMPLOS = {
    "objetos_contratuais": [
        ["Resíduos e Logística Reversa", "residuos-logistica-reversa",
         "Categoria mãe", "", 10],
        ["Pilhas e Baterias", "pilhas-e-baterias",
         "Pilhas, baterias e acumuladores elétricos", "residuos-logistica-reversa", 1],
    ],
    "normas_legais": [
        ["Lei 14.133/2021",
         "Lei nº 14.133, de 1º de abril de 2021 — Nova Lei de Licitações",
         "lei-14133-2021", "LEI_FEDERAL", "FEDERAL", "14.133/2021",
         "2021-04-01", "https://www.planalto.gov.br/ccivil_03/_ato2019-2022/2021/lei/l14133.htm",
         "Lei de Licitações e Contratos Administrativos."],
    ],
    "selos_certificacao": [
        ["INMETRO", "inmetro", "NACIONAL",
         "Instituto Nacional de Metrologia, Qualidade e Tecnologia",
         "https://www.gov.br/inmetro/"],
    ],
    "tags": [
        ["logística-reversa", "logistica-reversa"],
    ],
    "criterios": [
        ["CRIT-2026-001", "Pilhas alcalinas com logística reversa", "",
         "pilhas-e-baterias", "BENS", "TR_PB", "BASICO",
         "Exigir do fornecedor a apresentação de comprovante de adesão a sistema de logística reversa.",
         "Lei 12.305/2010, art. 33; Decreto 10.936/2022.",
         "Reduz contaminação do solo e impacto ambiental decorrente do descarte irregular.",
         "Apresentação semestral de comprovante de coleta junto a operador licenciado.",
         "",
         "12;15",
         "lei-14133-2021",
         "inmetro",
         "logistica-reversa",
         "Guia AGU 2024 — item 27",
         "RASCUNHO"],
    ],
    "paginas_institucionais": [],
    "glossario": [
        ["Logística Reversa", "logistica-reversa", "LR",
         "Conjunto de ações para coletar e dar destinação ambientalmente adequada aos produtos pós-consumo."],
        ["Estudo Técnico Preliminar", "estudo-tecnico-preliminar", "ETP",
         "Documento que precede a contratação, com análise de viabilidade e dimensionamento (Lei 14.133/2021)."],
    ],
}


HEADER_FILL = PatternFill("solid", fgColor="ED1C24")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)


def gerar(destino: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for aba, headers in HEADERS.items():
        ws = wb.create_sheet(title=aba)
        ws.append(headers)
        # Estilo do cabeçalho
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(headers[col_idx - 1]) + 4)
        # Exemplos (linhas de demonstração)
        for linha in EXEMPLOS.get(aba, []):
            ws.append(linha)
        ws.freeze_panes = "A2"

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(destino))
    print(f"Planilha-modelo gerada em: {destino}")


if __name__ == "__main__":
    raiz = Path(__file__).resolve().parent.parent
    destino = raiz / "seeds" / "pescp_seed_template.xlsx"
    gerar(destino)
