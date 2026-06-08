"""Importador idempotente de planilha XLSX para a PESCP.

Uso:
    python manage.py import_pescp_seed --file=seeds/pescp_seed_v1.xlsx
    python manage.py import_pescp_seed --file=seeds/pescp_seed_v1.xlsx --dry-run
    python manage.py import_pescp_seed --file=seeds/pescp_seed_v1.xlsx --update-existing

Estrutura esperada da planilha (uma aba por entidade — ver
docs/IMPORTACAO_SEED.md para detalhes completos):

- objetos_contratuais
- normas_legais
- selos_certificacao
- tags
- criterios
- paginas_institucionais
- glossario

Ordem de processamento:
    Ods -> NormaLegal -> SeloCertificacao -> Tag -> ObjetoContratual
       -> Criterio -> PaginaInstitucional -> TermoGlossario

Os 17 ODS oficiais são criados pelo comando ``create_initial_ods`` —
este comando NÃO recria ODS (assume que ``create_initial_ods`` já rodou).
"""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterator

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from apps.conteudo.models import PaginaInstitucional, TermoGlossario
from apps.core.models import Tag
from apps.criterios.models import (
    Criterio,
    FaseProcesso,
    NivelAmbicao,
    ObjetoContratual,
    StatusCriterio,
    TipoContratacao,
)
from apps.normas.models import NormaLegal, Ods, SeloCertificacao

log = logging.getLogger("apps.seed.import_pescp_seed")


# -----------------------------------------------------------------------------
# Helpers de parsing
# -----------------------------------------------------------------------------
def _normalizar(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _para_data(v: Any) -> date | None:
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _normalizar(v)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _para_int(v: Any) -> int | None:
    if v in (None, ""):
        return None
    try:
        return int(v)
    except (TypeError, ValueError):
        return None


def _para_bool(v: Any) -> bool:
    s = _normalizar(v).lower()
    return s in ("1", "true", "verdadeiro", "sim", "yes", "x")


def _split_semicolons(v: Any) -> list[str]:
    s = _normalizar(v)
    if not s:
        return []
    return [p.strip() for p in s.split(";") if p.strip()]


def _linhas(ws: Worksheet) -> Iterator[tuple[int, dict[str, Any]]]:
    """Itera (numero_linha_1based, dict header->valor) da aba.

    A primeira linha é o cabeçalho. Linhas totalmente vazias são puladas.
    """
    rows = ws.iter_rows(values_only=True)
    headers = [_normalizar(h).lower() for h in next(rows, [])]
    for idx, row in enumerate(rows, start=2):
        if not any(c not in (None, "") for c in row):
            continue
        yield idx, {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}


# -----------------------------------------------------------------------------
# Importadores por aba — cada um devolve (criados, atualizados, erros)
# -----------------------------------------------------------------------------
class ResultadoAba:
    def __init__(self, nome: str):
        self.nome = nome
        self.criados = 0
        self.atualizados = 0
        self.erros: list[str] = []
        self.pulados = 0

    def __str__(self) -> str:
        return (
            f"  [{self.nome}] criados={self.criados} atualizados={self.atualizados} "
            f"pulados={self.pulados} erros={len(self.erros)}"
        )


def _import_objetos(ws: Worksheet, update_existing: bool) -> ResultadoAba:
    r = ResultadoAba("objetos_contratuais")
    pendentes_parent: list[tuple[ObjetoContratual, str]] = []

    for linha, row in _linhas(ws):
        nome = _normalizar(row.get("nome"))
        slug = _normalizar(row.get("slug"))
        if not nome or not slug:
            r.erros.append(f"linha {linha}: 'nome' e 'slug' são obrigatórios")
            continue
        defaults = {
            "nome": nome,
            "descricao": _normalizar(row.get("descricao")),
            "ordem": _para_int(row.get("ordem")) or 0,
            "ativo": True,
        }
        obj, created = ObjetoContratual.objects.get_or_create(slug=slug, defaults=defaults)
        if created:
            r.criados += 1
        else:
            if update_existing:
                for k, v in defaults.items():
                    setattr(obj, k, v)
                obj.save()
                r.atualizados += 1
            else:
                r.pulados += 1
        # Parent é resolvido após primeira passada (pode referenciar slug
        # ainda não criado dentro da mesma aba)
        parent_slug = _normalizar(row.get("supercategoria_slug"))
        if parent_slug:
            pendentes_parent.append((obj, parent_slug))

    for obj, parent_slug in pendentes_parent:
        try:
            parent = ObjetoContratual.objects.get(slug=parent_slug)
            if obj.parent_id != parent.pk:
                obj.parent = parent
                obj.save(update_fields=["parent"])
        except ObjetoContratual.DoesNotExist:
            r.erros.append(f"objeto '{obj.slug}': supercategoria_slug='{parent_slug}' não existe")

    return r


def _import_normas(ws: Worksheet, update_existing: bool) -> ResultadoAba:
    r = ResultadoAba("normas_legais")
    for linha, row in _linhas(ws):
        slug = _normalizar(row.get("slug"))
        titulo_curto = _normalizar(row.get("titulo_curto"))
        if not slug or not titulo_curto:
            r.erros.append(f"linha {linha}: 'slug' e 'titulo_curto' são obrigatórios")
            continue
        defaults = {
            "titulo_curto": titulo_curto,
            "titulo_completo": _normalizar(row.get("titulo_completo")) or titulo_curto,
            "tipo": _normalizar(row.get("tipo")) or NormaLegal.Tipo.LEI_FEDERAL,
            "esfera": _normalizar(row.get("esfera")) or NormaLegal.Esfera.FEDERAL,
            "numero": _normalizar(row.get("numero")),
            "data_publicacao": _para_data(row.get("data_publicacao")),
            "link_oficial": _normalizar(row.get("link_oficial")),
            "ementa": _normalizar(row.get("ementa")),
        }
        obj, created = NormaLegal.objects.get_or_create(slug=slug, defaults=defaults)
        if created:
            r.criados += 1
        elif update_existing:
            for k, v in defaults.items():
                setattr(obj, k, v)
            obj.save()
            r.atualizados += 1
        else:
            r.pulados += 1
    return r


def _import_selos(ws: Worksheet, update_existing: bool) -> ResultadoAba:
    r = ResultadoAba("selos_certificacao")
    for linha, row in _linhas(ws):
        slug = _normalizar(row.get("slug"))
        nome = _normalizar(row.get("nome"))
        if not slug or not nome:
            r.erros.append(f"linha {linha}: 'slug' e 'nome' são obrigatórios")
            continue
        defaults = {
            "nome": nome,
            "tipo": _normalizar(row.get("tipo")) or SeloCertificacao.Tipo.NACIONAL,
            "descricao": _normalizar(row.get("descricao")),
            "link_oficial": _normalizar(row.get("link_oficial")),
        }
        obj, created = SeloCertificacao.objects.get_or_create(slug=slug, defaults=defaults)
        if created:
            r.criados += 1
        elif update_existing:
            for k, v in defaults.items():
                setattr(obj, k, v)
            obj.save()
            r.atualizados += 1
        else:
            r.pulados += 1
    return r


def _import_tags(ws: Worksheet, update_existing: bool) -> ResultadoAba:
    r = ResultadoAba("tags")
    for linha, row in _linhas(ws):
        slug = _normalizar(row.get("slug"))
        nome = _normalizar(row.get("nome"))
        if not slug or not nome:
            r.erros.append(f"linha {linha}: 'slug' e 'nome' são obrigatórios")
            continue
        obj, created = Tag.objects.get_or_create(slug=slug, defaults={"nome": nome})
        if created:
            r.criados += 1
        elif update_existing and obj.nome != nome:
            obj.nome = nome
            obj.save()
            r.atualizados += 1
        else:
            r.pulados += 1
    return r


def _import_criterios(ws: Worksheet, update_existing: bool) -> ResultadoAba:
    r = ResultadoAba("criterios")
    for linha, row in _linhas(ws):
        codigo = _normalizar(row.get("codigo"))
        titulo = _normalizar(row.get("titulo"))
        objeto_slug = _normalizar(row.get("objeto_contratual_slug"))
        if not codigo or not titulo or not objeto_slug:
            r.erros.append(
                f"linha {linha}: 'codigo', 'titulo' e 'objeto_contratual_slug' são obrigatórios"
            )
            continue
        try:
            objeto = ObjetoContratual.objects.get(slug=objeto_slug)
        except ObjetoContratual.DoesNotExist:
            r.erros.append(f"linha {linha}: objeto_contratual_slug='{objeto_slug}' não existe")
            continue

        defaults = {
            "titulo": titulo,
            "objeto_contratual": objeto,
            "tipo_contratacao": _normalizar(row.get("tipo_contratacao")) or TipoContratacao.BENS,
            "fase_processo": _normalizar(row.get("fase_processo")) or FaseProcesso.TR_PB,
            "nivel_ambicao": _normalizar(row.get("nivel_ambicao")) or NivelAmbicao.BASICO,
            "texto_providencia": _normalizar(row.get("texto_providencia")),
            "determinacoes_principais": _normalizar(row.get("determinacoes_principais")),
            "justificativa_tecnica": _normalizar(row.get("justificativa_tecnica")),
            "metodo_verificacao": _normalizar(row.get("metodo_verificacao")),
            "precaucoes": _normalizar(row.get("precaucoes")),
            "fonte": _normalizar(row.get("fonte")),
            "status": _normalizar(row.get("status")) or StatusCriterio.RASCUNHO,
        }
        slug = _normalizar(row.get("slug"))
        if slug:
            defaults["slug"] = slug
        obj, created = Criterio.objects.get_or_create(codigo=codigo, defaults=defaults)
        if not created and update_existing:
            for k, v in defaults.items():
                setattr(obj, k, v)
            obj.save()
        if created:
            r.criados += 1
        elif update_existing:
            r.atualizados += 1
        else:
            r.pulados += 1
            continue

        # Relacionamentos M2M
        # ODS — números separados por ;
        ods_nums = [int(n) for n in _split_semicolons(row.get("ods_numeros")) if n.isdigit()]
        if ods_nums:
            ods_qs = Ods.objects.filter(numero__in=ods_nums)
            obj.ods.set(ods_qs)
            faltantes = set(ods_nums) - set(ods_qs.values_list("numero", flat=True))
            if faltantes:
                r.erros.append(f"critério '{codigo}': ODS inexistentes {sorted(faltantes)}")

        normas_slugs = _split_semicolons(row.get("normas_slugs"))
        if normas_slugs:
            normas_qs = NormaLegal.objects.filter(slug__in=normas_slugs)
            obj.normas.set(normas_qs)
            faltantes = set(normas_slugs) - set(normas_qs.values_list("slug", flat=True))
            if faltantes:
                r.erros.append(f"critério '{codigo}': normas inexistentes {sorted(faltantes)}")

        selos_slugs = _split_semicolons(row.get("selos_slugs"))
        if selos_slugs:
            selos_qs = SeloCertificacao.objects.filter(slug__in=selos_slugs)
            obj.selos.set(selos_qs)

        tags_slugs = _split_semicolons(row.get("tags_slugs"))
        if tags_slugs:
            tags_qs = Tag.objects.filter(slug__in=tags_slugs)
            obj.tags.set(tags_qs)
    return r


def _import_paginas(ws: Worksheet, update_existing: bool) -> ResultadoAba:
    r = ResultadoAba("paginas_institucionais")
    for linha, row in _linhas(ws):
        slug = _normalizar(row.get("slug"))
        titulo = _normalizar(row.get("titulo"))
        if not slug or not titulo:
            r.erros.append(f"linha {linha}: 'slug' e 'titulo' são obrigatórios")
            continue
        defaults = {
            "titulo": titulo,
            "conteudo": _normalizar(row.get("conteudo")),
            "ordem": _para_int(row.get("ordem")) or 0,
            "publicada": _para_bool(row.get("publicada")) if row.get("publicada") is not None else True,
        }
        obj, created = PaginaInstitucional.objects.get_or_create(slug=slug, defaults=defaults)
        if created:
            r.criados += 1
        elif update_existing:
            for k, v in defaults.items():
                setattr(obj, k, v)
            obj.save()
            r.atualizados += 1
        else:
            r.pulados += 1
    return r


def _import_glossario(ws: Worksheet, update_existing: bool) -> ResultadoAba:
    r = ResultadoAba("glossario")
    for linha, row in _linhas(ws):
        slug = _normalizar(row.get("slug"))
        termo = _normalizar(row.get("termo"))
        if not slug or not termo:
            r.erros.append(f"linha {linha}: 'slug' e 'termo' são obrigatórios")
            continue
        defaults = {
            "termo": termo,
            "definicao": _normalizar(row.get("definicao")),
            "sigla": _normalizar(row.get("sigla")),
        }
        obj, created = TermoGlossario.objects.get_or_create(slug=slug, defaults=defaults)
        if created:
            r.criados += 1
        elif update_existing:
            for k, v in defaults.items():
                setattr(obj, k, v)
            obj.save()
            r.atualizados += 1
        else:
            r.pulados += 1
    return r


# -----------------------------------------------------------------------------
# Command
# -----------------------------------------------------------------------------
HANDLERS = [
    ("objetos_contratuais", _import_objetos),
    ("normas_legais", _import_normas),
    ("selos_certificacao", _import_selos),
    ("tags", _import_tags),
    ("criterios", _import_criterios),
    ("paginas_institucionais", _import_paginas),
    ("glossario", _import_glossario),
]


class Command(BaseCommand):
    help = "Importa planilha XLSX de seed da PESCP (idempotente, transacional)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--file",
            required=True,
            help="Caminho para o arquivo .xlsx",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Executa em transação e dá rollback ao final — útil para validar planilha.",
        )
        parser.add_argument(
            "--update-existing",
            action="store_true",
            help="Atualiza registros existentes (default: apenas cria novos).",
        )

    def handle(self, *args, **options):
        path = Path(options["file"])
        if not path.exists():
            raise CommandError(f"Arquivo não encontrado: {path}")

        dry_run: bool = options["dry_run"]
        update_existing: bool = options["update_existing"]

        self.stdout.write(self.style.MIGRATE_HEADING(f"\nPESCP — importando {path}"))
        if dry_run:
            self.stdout.write(self.style.WARNING("  modo dry-run: tudo é revertido ao final"))
        if update_existing:
            self.stdout.write("  modo update-existing: atualiza registros já existentes")
        self.stdout.write("")

        wb = load_workbook(filename=str(path), read_only=True, data_only=True)

        resultados: list[ResultadoAba] = []
        try:
            with transaction.atomic():
                for nome_aba, handler in HANDLERS:
                    if nome_aba not in wb.sheetnames:
                        self.stdout.write(self.style.WARNING(f"  aba ausente: {nome_aba} (pulando)"))
                        continue
                    ws = wb[nome_aba]
                    resultado = handler(ws, update_existing=update_existing)
                    resultados.append(resultado)
                    self.stdout.write(str(resultado))
                    for err in resultado.erros:
                        self.stdout.write(self.style.ERROR(f"    {err}"))

                if dry_run:
                    raise _DryRunRollback("dry-run: rollback forçado")
        except _DryRunRollback:
            self.stdout.write(self.style.WARNING("\nDry-run finalizado — nenhuma alteração persistida."))
            return

        total_criados = sum(r.criados for r in resultados)
        total_atualizados = sum(r.atualizados for r in resultados)
        total_erros = sum(len(r.erros) for r in resultados)

        self.stdout.write("")
        self.stdout.write(
            self.style.SUCCESS(
                f"Importação concluída — criados={total_criados} atualizados={total_atualizados} "
                f"erros={total_erros}"
            )
        )


class _DryRunRollback(Exception):
    """Exceção interna para forçar rollback em dry-run."""
