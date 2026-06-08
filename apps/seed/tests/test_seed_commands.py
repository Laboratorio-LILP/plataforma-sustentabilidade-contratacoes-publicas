"""Testes dos management commands de seed."""

import pytest
from django.core.management import call_command

from apps.conteudo.models import PaginaInstitucional
from apps.normas.models import Ods


pytestmark = pytest.mark.django_db


def test_create_initial_ods_cria_17(db):
    call_command("create_initial_ods")
    assert Ods.objects.count() == 17
    ods7 = Ods.objects.get(numero=7)
    assert "Energia" in ods7.nome


def test_create_initial_ods_idempotente(db):
    call_command("create_initial_ods")
    call_command("create_initial_ods")
    assert Ods.objects.count() == 17


def test_create_initial_pages_cria_paginas_esperadas(db):
    call_command("create_initial_pages")
    slugs = set(PaginaInstitucional.objects.values_list("slug", flat=True))
    esperados = {
        "sobre", "governanca", "metodologia", "parceiros", "contato",
        "transparencia", "acessibilidade",
        "politica-de-privacidade", "politica-de-cookies",
        "mapa-do-site", "fale-conosco",
    }
    assert esperados.issubset(slugs)


def test_import_pescp_seed_xlsx_minimo(tmp_path, db):
    """Cria planilha mínima em memória, importa, valida contagens."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    # Aba: objetos_contratuais
    ws = wb.create_sheet("objetos_contratuais")
    ws.append(["nome", "slug", "descricao", "supercategoria_slug", "ordem"])
    ws.append(["Resíduos", "residuos", "Super", "", 1])
    ws.append(["Pilhas e Baterias", "pilhas-e-baterias", "Cat", "residuos", 1])

    # Aba: normas_legais
    ws = wb.create_sheet("normas_legais")
    ws.append(["titulo_curto", "titulo_completo", "slug", "tipo", "esfera", "numero", "data_publicacao", "link_oficial", "ementa"])
    ws.append(["Lei 14.133/2021", "Lei nº 14.133...", "lei-14133-2021", "LEI_FEDERAL", "FEDERAL", "14.133/2021", "", "", ""])

    # Aba: selos_certificacao
    ws = wb.create_sheet("selos_certificacao")
    ws.append(["nome", "slug", "tipo", "descricao", "link_oficial"])
    ws.append(["INMETRO", "inmetro", "NACIONAL", "", ""])

    # Aba: tags
    ws = wb.create_sheet("tags")
    ws.append(["nome", "slug"])
    ws.append(["logística-reversa", "logistica-reversa"])

    # Aba: criterios
    ws = wb.create_sheet("criterios")
    ws.append([
        "codigo", "titulo", "slug", "objeto_contratual_slug",
        "tipo_contratacao", "fase_processo", "nivel_ambicao",
        "texto_providencia", "determinacoes_principais",
        "justificativa_tecnica", "metodo_verificacao", "precaucoes",
        "ods_numeros", "normas_slugs", "selos_slugs", "tags_slugs",
        "fonte", "status",
    ])
    ws.append([
        "CRIT-2026-100", "Pilhas alcalinas com logística reversa", "",
        "pilhas-e-baterias", "BENS", "TR_PB", "BASICO",
        "<p>Exigir LRR.</p>", "<p>PNRS.</p>", "<p>Reduz contaminação.</p>",
        "<p>Comprovante de coleta.</p>", "",
        "12;15", "lei-14133-2021", "inmetro", "logistica-reversa",
        "Guia AGU 2024", "PUBLICADO",
    ])

    # Aba: paginas_institucionais (vazia, mas presente)
    ws = wb.create_sheet("paginas_institucionais")
    ws.append(["titulo", "slug", "conteudo", "ordem", "publicada"])

    # Aba: glossario
    ws = wb.create_sheet("glossario")
    ws.append(["termo", "slug", "sigla", "definicao"])
    ws.append(["Logística Reversa", "logistica-reversa", "LR", "Coleta pós-consumo."])

    path = tmp_path / "seed.xlsx"
    wb.save(str(path))

    # Precondição: ODS precisam existir (para ods_numeros)
    call_command("create_initial_ods")

    # Importa
    call_command("import_pescp_seed", "--file", str(path))

    from apps.conteudo.models import TermoGlossario
    from apps.core.models import Tag
    from apps.criterios.models import Criterio, ObjetoContratual
    from apps.normas.models import NormaLegal, SeloCertificacao

    assert ObjetoContratual.objects.count() == 2
    assert ObjetoContratual.objects.get(slug="pilhas-e-baterias").parent.slug == "residuos"
    assert NormaLegal.objects.filter(slug="lei-14133-2021").exists()
    assert SeloCertificacao.objects.filter(slug="inmetro").exists()
    assert Tag.objects.filter(slug="logistica-reversa").exists()
    c = Criterio.objects.get(codigo="CRIT-2026-100")
    assert c.objeto_contratual.slug == "pilhas-e-baterias"
    assert set(c.ods.values_list("numero", flat=True)) == {12, 15}
    assert TermoGlossario.objects.filter(slug="logistica-reversa").exists()


def test_import_pescp_seed_respeita_slug_da_planilha(tmp_path, db):
    """Slug fornecido na planilha de critérios deve ser preservado;
    quando vazio, o save() do model auto-gera."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("objetos_contratuais")
    ws.append(["nome", "slug", "descricao", "supercategoria_slug", "ordem"])
    ws.append(["Pilhas", "pilhas", "", "", 1])

    ws = wb.create_sheet("criterios")
    ws.append([
        "codigo", "titulo", "slug", "objeto_contratual_slug",
        "tipo_contratacao", "fase_processo", "nivel_ambicao",
        "texto_providencia", "determinacoes_principais",
        "justificativa_tecnica", "metodo_verificacao", "precaucoes",
        "ods_numeros", "normas_slugs", "selos_slugs", "tags_slugs",
        "fonte", "status",
    ])
    ws.append([
        "CRIT-2026-200", "Limites de metais pesados", "limites-metais-pesados",
        "pilhas", "BENS", "TR_PB", "BASICO",
        "", "", "", "", "",
        "", "", "", "",
        "", "PUBLICADO",
    ])
    ws.append([
        "CRIT-2026-201", "Critério sem slug explícito", "",
        "pilhas", "BENS", "TR_PB", "BASICO",
        "", "", "", "", "",
        "", "", "", "",
        "", "PUBLICADO",
    ])

    path = tmp_path / "seed.xlsx"
    wb.save(str(path))

    call_command("create_initial_ods")
    call_command("import_pescp_seed", "--file", str(path))

    from apps.criterios.models import Criterio

    c1 = Criterio.objects.get(codigo="CRIT-2026-200")
    assert c1.slug == "limites-metais-pesados"
    c2 = Criterio.objects.get(codigo="CRIT-2026-201")
    assert c2.slug.startswith("crit-2026-201-")


def test_import_pescp_seed_idempotente(tmp_path, db):
    """Rodar o importador duas vezes não cria duplicatas."""
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("objetos_contratuais")
    ws.append(["nome", "slug", "descricao", "supercategoria_slug", "ordem"])
    ws.append(["Veículos", "veiculos", "", "", 1])
    path = tmp_path / "seed.xlsx"
    wb.save(str(path))

    call_command("import_pescp_seed", "--file", str(path))
    call_command("import_pescp_seed", "--file", str(path))

    from apps.criterios.models import ObjetoContratual

    assert ObjetoContratual.objects.filter(slug="veiculos").count() == 1
